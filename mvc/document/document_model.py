import os
import threading
from copy import copy
from pathlib import Path

import yaml
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5.QtCore import QDate, QObject, pyqtSignal


class DocumentModel(QObject):
    """Manages data and business logic for the document generation window.

    This class loads necessary configuration, filters materials based on
    document type, and handles the creation of styled Excel documents in a
    separate thread to keep the UI responsive.

    Attributes:
        show_notification: Signal emitting messages for the user.
        progress_changed: Signal to update the progress bar during export.
    """

    show_notification = pyqtSignal(str, str)
    progress_changed = pyqtSignal(str, int)

    def __init__(
        self,
        product_name: str,
        norms_calculations_value: int,
        materials: list[dict],
        current_product_path: str,
    ) -> None:
        """Initializes the DocumentModel.

        Args:
            product_name: The name of the product.
            norms_calculations_value: The quantity for which norms are calculated.
            materials: The full list of materials for the product.
            current_product_path: The filesystem path to the product's folder.
        """
        super().__init__()

        # Data from the main application window
        self.product_name: str = product_name
        self.materials: list[dict] = materials
        self.current_product_path: str = current_product_path

        # Data from configuration file
        self.signature_from_human: list[str] = []
        self.signature_from_position: list[str] = []
        self.signature_whom_human: list[str] = []
        self.signature_whom_position: list[str] = []
        self.document_blacklist: list[str] = []
        self.document_whitelist: list[str] = []
        self.bid_blacklist: list[str] = []
        self.bid_whitelist: list[str] = []
        self.templates_folder_path: str = ""
        self._load_config()

        # Data for document templates
        self.outgoing_number: str = ""
        self.current_date: str = ""
        self.quantity: int = norms_calculations_value
        self.whom_position: str = ""
        self.whom_fio: str = ""
        self.from_position: str = ""
        self.from_fio: str = ""

        # Progress bar settings
        self.progress_bar_export_excel_step_size: int = 7

    def get_current_date(self) -> QDate:
        """Returns the current date."""
        return QDate.currentDate()

    def get_desktop_path(self) -> Path | None:
        """Finds the path to the user's desktop folder.

        Checks for common OneDrive and standard desktop paths.

        Returns:
            A Path object to the desktop, or None if an error occurs.
        """
        try:
            home = Path.home()
            # Check for Russian OneDrive path
            onedrive_ru = home / "OneDrive" / "Рабочий стол"
            if onedrive_ru.exists():
                return onedrive_ru
            # Check for English OneDrive path
            onedrive_en = home / "OneDrive" / "Desktop"
            if onedrive_en.exists():
                return onedrive_en
            # Fallback to standard Desktop path
            return home / "Desktop"
        except Exception as e:
            self.show_notification.emit("error", f"Ошибка при получении пути к рабочему столу: {e}")
            return None

    def export_in_thread(
        self,
        document_type: str,
        save_folder_path: str,
    ) -> threading.Thread | None:
        """Starts the document export process in a separate thread.

        Args:
            document_type: The type of document to export ('document' or 'bid').
            save_folder_path: The folder where the document will be saved.

        Returns:
            The thread object that was started, or None on error.
        """
        path = save_folder_path
        if not path:
            desktop = self.get_desktop_path()
            if not desktop:
                self.show_notification.emit("error", "Не удалось найти путь к рабочему столу.")
                return None
            path = str(desktop)

        try:
            thread = threading.Thread(
                target=self._export_to_excel, args=(document_type, path)
            )
            thread.daemon = True
            thread.start()
            return thread
        except Exception as e:
            self.show_notification.emit(
                "error", f"Не удалось запустить поток для экспорта: {e}"
            )
            return None

    def _load_config(self) -> None:
        """Loads configuration from the config.yaml file."""
        try:
            with open("config.yaml", "r", encoding="utf-8") as file:
                config = yaml.safe_load(file) or {{}}
        except FileNotFoundError:
            self.show_notification.emit("error", "Файл config.yaml не найден.")
            return
        
        # Load templates folder path with fallback to local ./templates
        configured_templates = config.get("templates_folder_path")
        if configured_templates and os.path.isdir(configured_templates):
            self.templates_folder_path = configured_templates
        else:
            local_templates = os.path.join(os.getcwd(), "templates")
            if os.path.isdir(local_templates):
                self.templates_folder_path = local_templates
            else:
                self.templates_folder_path = ""
                self.show_notification.emit(
                    "error",
                    "Путь к папке templates не найден. Укажите templates_folder_path в config.yaml.",
                )

        # Load signatures
        self.signature_from_human = config.get("signature_from_human", [])
        self.signature_from_position = config.get("signature_from_position", [])
        self.signature_whom_human = config.get("signature_whom_human", [])
        self.signature_whom_position = config.get("signature_whom_position", [])

        # Load whitelists and blacklists
        self.bid_blacklist = config.get("bid_blacklist", [])
        self.document_blacklist = config.get("document_blacklist", [])
        self.bid_whitelist = config.get("bid_whitelist", [])
        self.document_whitelist = config.get("document_whitelist", [])

    def _get_document_materials_list(self) -> list[dict]:
        """Filters and returns the material list for a formal note ('document').

        Filtering logic:
        1.  Always include materials in the document_whitelist.
        2.  For other materials, include them if their unit is NOT 'pcs' and
            they are not from an RMP folder, unless they are in the
            document_blacklist.

        Returns:
            A list of filtered materials.
        """
        filtered_materials = []
        try:
            for item in self.materials:
                nomenclature_lower = item["Номенклатура"].lower()

                # 1. Whitelist check
                if any(
                    word.lower() in nomenclature_lower
                    for word in self.document_whitelist
                    if word
                ):
                    filtered_materials.append(item)
                    continue

                # 2. Main filter for non-whitelisted items
                if item["Ед. изм."] != "шт" and not item["РМП"]:
                    is_in_blacklist = any(
                        word.lower() in nomenclature_lower
                        for word in self.document_blacklist
                        if word
                    )
                    if not is_in_blacklist:
                        filtered_materials.append(item)
            return filtered_materials
        except Exception as e:
            self.show_notification.emit(
                "error", f"Произошла ошибка при фильтрации материалов:\n{e}"
            )
            return []

    def _get_bid_materials_list(self) -> list[dict]:
        """Filters and returns the material list for a request ('bid').

        Filtering logic:
        1.  Always include materials in the bid_whitelist.
        2.  For other materials, include them if their unit IS 'pcs', unless
            they are in the bid_blacklist.

        Returns:
            A list of filtered materials.
        """
        filtered_materials = []
        try:
            for item in self.materials:
                nomenclature_lower = item["Номенклатура"].lower()

                # 1. Whitelist check
                if any(
                    word.lower() in nomenclature_lower
                    for word in self.bid_whitelist
                    if word
                ):
                    filtered_materials.append(item)
                    continue

                # 2. Main filter for non-whitelisted items
                if item["Ед. изм."] == "шт":
                    is_in_blacklist = any(
                        word.lower() in nomenclature_lower
                        for word in self.bid_blacklist
                        if word
                    )
                    if not is_in_blacklist:
                        filtered_materials.append(item)
            return filtered_materials
        except Exception as e:
            self.show_notification.emit(
                "error", f"Произошла ошибка при фильтрации материалов:\n{e}"
            )
            return []

    def _export_materials_list(
        self,
        workbook: Workbook,
        materials_list: list[dict],
        progress_bar_value: int,
        progress_bar_process_text: str,
    ) -> int:
        """Creates and styles a 'Material List' sheet in the given workbook.

        Args:
            workbook: The openpyxl Workbook to add the sheet to.
            materials_list: The list of materials to populate the sheet with.
            progress_bar_value: The current value of the progress bar.
            progress_bar_process_text: The text for the progress bar.

        Returns:
            The updated progress bar value.
        """
        if not self.templates_folder_path:
            self.show_notification.emit(
                "error",
                "Путь к шаблонам не задан. Укажите templates_folder_path в config.yaml.",
            )
            return progress_bar_value

        try:
            table_template_path = os.path.join(self.templates_folder_path, "table.xlsx")
            if not os.path.exists(table_template_path):
                self.show_notification.emit(
                    "error",
                    f"Файл шаблона не найден: {table_template_path}",
                )
                return progress_bar_value
            template_wb = load_workbook(table_template_path)
            template_sheet = template_wb.active
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            new_sheet = workbook.create_sheet(title="Список материалов")
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Safely copy values and styles from the template sheet
            for row in template_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(
                        row=cell.row, column=cell.column, value=cell.value
                    )
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Copy column dimensions and row heights
            for col_letter, col_dim in template_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = col_dim.width
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            for row_idx, row_dim in template_sheet.row_dimensions.items():
                new_sheet.row_dimensions[row_idx].height = row_dim.height
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Fill in the material data
            start_row = 2
            for i, item in enumerate(materials_list):
                row = start_row + i
                new_sheet.cell(row=row, column=1, value=item["Номенклатура"])
                new_sheet.cell(row=row, column=2, value=item["Ед. изм."])
                new_sheet.cell(row=row, column=3, value=item["Количество"])
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Apply font and alignment
            font = Font(name="Times New Roman", size=14)
            for row in new_sheet.iter_rows(
                min_row=start_row, max_row=new_sheet.max_row, max_col=3
            ):
                for cell in row:
                    cell.font = font
                    if cell.column == 1:
                        cell.alignment = Alignment(
                            horizontal="left", vertical="top", wrap_text=True
                        )
                    else:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="top", wrap_text=True
                        )
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Apply borders
            thick_side = Side(border_style="thick", color="000000")
            thin_side = Side(border_style="thin", color="000000")
            for row_idx in range(start_row, new_sheet.max_row + 1):
                for col_idx in range(1, 4):
                    cell = new_sheet.cell(row=row_idx, column=col_idx)
                    left = thick_side if col_idx == 1 else thin_side
                    right = thick_side if col_idx == 3 else thin_side
                    cell.border = Border(
                        left=left, right=right, top=thin_side, bottom=thin_side
                    )
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Page setup
            new_sheet.page_setup.fitToWidth = 1
            new_sheet.page_setup.fitToHeight = 0
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            return progress_bar_value
        except Exception as e:
            self.show_notification.emit(
                "error", f"Ошибка при создании листа с перечнем материалов:\n{e}"
            )
            return progress_bar_value

    def _export_to_excel(
        self,
        document_type: str,
        save_folder_path: str,
    ) -> None:
        """Orchestrates the Excel export process for a given document type."""
        if not self.templates_folder_path:
            self.show_notification.emit(
                "error",
                "Путь к шаблонам не задан. Укажите templates_folder_path в config.yaml.",
            )
            return
        context = {
            "outgoing_number": self.outgoing_number,
            "current_date": self.current_date,
            "whom_position": self.whom_position,
            "whom_fio": self.whom_fio,
            "product_name": self.product_name,
            "product_quantity": self.quantity,
            "from_position": self.from_position,
            "from_fio": self.from_fio,
        }
        progress_bar_value = 0

        try:
            if document_type == "document":
                template_path = os.path.join(self.templates_folder_path, "document.xlsx")
                save_filename = f"Докладная записка {self.product_name}.xlsx"
                sheet_title = "Докладная записка"
                materials_list = self._get_document_materials_list()
            elif document_type == "bid":
                template_path = os.path.join(self.templates_folder_path, "bid.xlsx")
                save_filename = f"Заявка {self.product_name}.xlsx"
                sheet_title = "Заявка"
                materials_list = self._get_bid_materials_list()
            else:
                return  # Should not happen

            if not os.path.exists(template_path):
                self.show_notification.emit(
                    "error",
                    f"???? ??????? ?? ??????: {template_path}",
                )
                return

            save_path = Path(save_folder_path) / save_filename
            progress_text = f"Экспорт в {save_filename}..."

            # Load main template and render context variables
            wb = load_workbook(template_path)
            ws = wb.active
            ws.title = sheet_title
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_text, progress_bar_value)

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "{{" in cell.value:
                        template = Template(cell.value)
                        cell.value = template.render(context)
            progress_bar_value += self.progress_bar_export_excel_step_size
            self.progress_changed.emit(progress_text, progress_bar_value)

            # Create and append the materials list sheet
            progress_bar_value = self._export_materials_list(
                workbook=wb,
                materials_list=materials_list,
                progress_bar_value=progress_bar_value,
                progress_bar_process_text=progress_text,
            )

            # Save the final workbook
            wb.save(save_path)
            self.progress_changed.emit("Экспорт завершен", 100)
            self.show_notification.emit("info", f"Экспорт в {save_filename} успешно завершен")

        except Exception as e:
            self.progress_changed.emit("Экспорт не удался", 100)
            self.show_notification.emit(
                "error", f"Произошла ошибка при сохранении документа: {e}"
            )
