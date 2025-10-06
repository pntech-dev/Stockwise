import os
import yaml
import subprocess
import pandas as pd

from PyQt5.QtCore import QObject, pyqtSignal

# Экспорт в Word
from docx import Document

# Эксопрт в Excel
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Экспортв PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class Model(QObject):
    progress_changed = pyqtSignal(str, int) # Сигнал изменения значения прогресс бара
    show_notification = pyqtSignal(str, str) # Сигнал показа уведомления

    def __init__(self):
        super().__init__()

        self.program_version_number = None # Версия программы
        self.program_server_path = None # Путь к программе на сервере
        
        self.path_to_products_folder = None
        self.is_products_folder_available = False
        self.__load_config()  # Загружаем конфигурацию при инициализации

        self.products_names = []  # Список названий изделий
        self.update_product_names()  # Обновляем список названий изделий

        self.current_product = ""  # Текущее выбранное изделие
        self.current_product_materials = []  # Список материалов текущего изделия

        self.norms_calculations_value = 1000 # Значение на которое рассчитываются нормы изделия

        # Настройки для прогресс бара
        self.progress_bar_export_to_excel_step_size = 11 # Excel
        self.progress_bar_export_to_word_step_size = 6 # Word
        self.progress_bar_export_to_pdf_step_size = 5 # PDF

    def __load_config(self):
        """Функция загружает конфигурацию из файла config.yaml."""
        try:
            with open("config.yaml", "r", encoding="utf-8") as file:
                config = yaml.safe_load(file)
        except FileNotFoundError:
            self.show_notification.emit("error", "Файл config.yaml не найден.")
            return
        
        if "path_to_products_folder" not in config:
            self.show_notification.emit("error", "В файле config.yaml не указан путь к папке изделий.")
            return

        # Записываем данные из файла кофигурации
        self.program_version_number = config["program_version_number"]
        self.program_server_path = config["server_program_path"]
        path = config["path_to_products_folder"]
        if os.path.exists(path) and os.path.isdir(path):
            self.path_to_products_folder = path
            self.is_products_folder_available = True
        else:
            self.show_notification.emit("error", "Указанный путь к папке изделий не существует или не является папкой.")
            return

    def __export_to_excel(self, save_path, data):
        """Функция экспортирует данные в Excel файл."""
        try:
            progress_bar_value = 0
            progress_bar_process_text = f"Экспортируем данные в документ {os.path.basename(save_path)}"

            df = pd.DataFrame(data) # Преобразуем данные в DataFrame

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Преобразуем списки в строки, как в других функциях экспорта
            for col in df.columns:
                if df[col].apply(lambda x: isinstance(x, list)).any():
                    df[col] = df[col].apply(lambda x: ", ".join(map(str, x)) if isinstance(x, list) else x)

            wb = Workbook()
            ws = wb.active
            ws.title = "Данные"

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Записываем данные в лист
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Определяем стиль границ
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Применяем границы и выравнивание ко всем ячейкам
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    # Устанавливаем перенос текста и выравнивание
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)
            
            # Устанавливаем фиксированную ширину колонок, чтобы поместиться на A4
            # Ширины колонок подобраны для стандартного листа A4 в книжной ориентации
            ws.column_dimensions['A'].width = 35  # Номенклатура
            ws.column_dimensions['B'].width = 12  # Количество
            ws.column_dimensions['C'].width = 10  # Ед. изм.
            ws.column_dimensions['D'].width = 35  # Изделие

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Настройка параметров страницы
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paper_size = ws.PAPERSIZE_A4

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)
            
            # Устанавливаем область печати и параметры для вписывания в страницу
            ws.print_area = f'A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}'
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0 # 0 означает, что высота будет подстраиваться

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)
            
            wb.save(save_path) # Сохраняем файл

            progress_bar_value = 100
            self.progress_changed.emit("Экспортирование завершено", progress_bar_value)
            self.show_notification.emit("info", f"Данные экспортированы в {os.path.basename(save_path)}")
        except Exception as e:
            self.show_notification.emit("error", f"Ошибка при экспорте в {os.path.basename(save_path)}: {e}")

    def __export_to_word(self, save_path, data):
        """Функция экспортирует данные в Word файл."""
        try:
            progress_bar_value = 0
            progress_bar_process_text = f"Экспортируем данные в документ {os.path.basename(save_path)}"

            document = Document() # Создаём экземпляр документа

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            df = pd.DataFrame(data) # Преобразуем данные в DataFrame

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)
            
            # Создаём и настраиваем таблицу
            table = document.add_table(rows=1, cols=len(df.columns))
            table.style = 'Table Grid'

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Определяем и записываем заголовки
            hdr_cells = table.rows[0].cells
            for i, col_name in enumerate(df.columns):
                hdr_cells[i].text = str(col_name)
            
            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Вносим данные в документ
            for index, row in df.iterrows():
                row_cells = table.add_row().cells
                for i, cell_value in enumerate(row):
                    if isinstance(cell_value, list):
                        cell_value = ", ".join(map(str, cell_value))
                    row_cells[i].text = str(cell_value)

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            document.save(save_path) # Сохраняем файл

            progress_bar_value = 100
            self.progress_changed.emit("Экспортирование завершено", progress_bar_value)
            self.show_notification.emit("info", f"Данные экспортированы в {os.path.basename(save_path)}")
        except Exception as e:
            self.show_notification.emit("error", f"Ошибка при экспорте в {os.path.basename(save_path)}: {e}")

    def __export_to_pdf(self, save_path, data):
        """Функция экспортирует данные в PDF файл с использованием ReportLab."""
        try:
            if not data:
                return
            
            progress_bar_value = 0
            progress_bar_process_text = f"Экспортируем данные в документ {os.path.basename(save_path)}"

            # --- Регистрация шрифтов --- 
            font_dir = os.path.join(os.environ.get("SystemRoot", "C:\\Windows"), "Fonts")
            pdfmetrics.registerFont(TTFont('Arial', os.path.join(font_dir, "Arial.ttf")))
            pdfmetrics.registerFont(TTFont('Arial-Bold', os.path.join(font_dir, "Arialbd.ttf")))
            
            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # Используем книжную ориентацию (letter)
            doc = SimpleDocTemplate(save_path, pagesize=letter)
            story = []
            
            styles = getSampleStyleSheet()
            paragraph_style = styles['Normal']
            paragraph_style.fontName = 'Arial'
            paragraph_style.fontSize = 8

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # --- Подготовка данных ---
            df = pd.DataFrame(data)
            for col in df.columns:
                if df[col].apply(lambda x: isinstance(x, list)).any():
                    df[col] = df[col].apply(lambda x: ", ".join(map(str, x)) if isinstance(x, list) else x)
            
            headings = list(df.columns)
            data_rows = []
            for index, row in df.iterrows():
                new_row = []
                for col_name in df.columns:
                    cell_data = str(row[col_name])
                    if col_name in ['Номенклатура', 'Изделие']:
                        new_row.append(Paragraph(cell_data, paragraph_style))
                    else:
                        new_row.append(cell_data)
                data_rows.append(new_row)

            table_data = [headings] + data_rows

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # --- Создание и стилизация таблицы ---
            # Ширина колонок для книжной ориентации
            col_widths = [158, 70, 50, 190]
            table = Table(table_data, colWidths=col_widths)
            
            # Черно-белый стиль
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Arial-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            
            table.setStyle(style)

            progress_bar_value += self.progress_bar_export_to_excel_step_size
            self.progress_changed.emit(progress_bar_process_text, progress_bar_value)

            # --- Сборка документа ---
            story.append(table)
            doc.build(story)

            progress_bar_value = 100
            self.progress_changed.emit("Экспортирование завершено", progress_bar_value)
            self.show_notification.emit("info", f"Данные экспортированы в {os.path.basename(save_path)}")
        except Exception as e:
            self.show_notification.emit("error", f"Ошибка при экспорте в {os.path.basename(save_path)}: {e}")

    def get_semi_finished_products(self, product_name):
        """Функция возвращает список полуфабрикатов входящих в продукт."""
        try:
            semi_finished_products = []  # Создаем пустой список полуфабрикатов
            # Формируем путь к папке изделия
            self.product_path = os.path.join(
                self.path_to_products_folder, *product_name)

            # Если путь к папке изделия существует
            if os.path.exists(self.product_path) and os.path.isdir(self.product_path):
                # Получаем список файлов и папок в папке изделия
                for item_name in os.listdir(self.product_path):
                    item_path = os.path.join(
                        self.product_path, item_name)  # Формируем путь к элементу

                    # Если это папка "рмп", обрабатываем ее содержимое
                    if item_name.lower() == "рмп" and os.path.isdir(item_path):
                        # Получаем список файлов в папке "рмп"
                        for rmp_file_name in os.listdir(item_path):
                            rmp_file_path = os.path.join(item_path, rmp_file_name)
                            # Проверяем, что это файл Excel
                            if os.path.isfile(rmp_file_path) and rmp_file_path.lower().endswith(('.xlsx', '.xls')):
                                semi_finished_products.append(rmp_file_path)
                    # Иначе, если это файл Excel
                    elif os.path.isfile(item_path) and item_path.lower().endswith(('.xlsx', '.xls')):
                        semi_finished_products.append(item_path)

            return semi_finished_products
        except:
            self.show_notification.emit("error", "Ошибка при получении списка полуфабрикатов.")

    def get_product_materials(self, semi_finished_products):
        """Функция возвращает список материалов входящих в продукт."""
        product_materials_dict = {}  # Используем словарь для быстрой проверки существования

        for semi_finished_product in semi_finished_products:
            if os.path.exists(semi_finished_product):  # Если полуфабрикат существует
                try:
                    # Определяем колонки для чтения
                    columns_to_read = ['Номенклатура', 'Количество', 'Ед. изм.']
                    # Читаем только нужные колонки из Excel файла
                    df = pd.read_excel(semi_finished_product, sheet_name='TDSheet', usecols=columns_to_read)
                    df = df.fillna('')  # Заменяем все NaN значения на пустые строки

                    # Преобразуем DataFrame в список словарей
                    data_list = df.to_dict('records')

                    for new_item in data_list:
                        # Рассчитываем норму материалов на заданное количество
                        if self.norms_calculations_value is not None and self.norms_calculations_value != 1000:
                            new_item['Количество'] = (new_item['Количество'] / 1000) * self.norms_calculations_value

                        nomenclature = new_item['Номенклатура']
                        product_name = os.path.splitext(
                            os.path.basename(semi_finished_product))[0]

                        # Если материал уже есть в словаре
                        if nomenclature in product_materials_dict:
                            existing_item = product_materials_dict[nomenclature]
                            # Увеличиваем количество
                            existing_item['Количество'] += new_item['Количество']

                            # Если название изделия не найдено в списке изделий
                            if product_name not in existing_item['Изделие']:
                                # Добавляем название изделия в список изделий
                                existing_item['Изделие'].append(product_name)
                        else:
                            # Если материал не найден, добавляем его в словарь
                            new_item['Изделие'] = [product_name]
                            product_materials_dict[nomenclature] = new_item
                except Exception as e:
                    self.show_notification.emit("error", f"Ошибка при чтении файла {semi_finished_product}: {e}")
            else:
                self.show_notification.emit("error", f"Файл {semi_finished_product} не найден.")
        return list(product_materials_dict.values())
    
    def get_desktop_path(self):
        """Функция возвращает путь к папке Desktop."""
        try:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            return desktop_path
        except Exception as e:
            self.show_notification.emit("error", f"Ошибка при получении пути к папке Desktop: {e}")

    def check_program_version(self):
        """Функция проверяет версию программы"""
        if not self.program_server_path or not os.path.exists(self.program_server_path):
            return None
        
        program_server_files = os.listdir(self.program_server_path)
        for file_name in program_server_files:
            # Ищем файл, который начинается с "config" и заканчивается ".yaml"
            if file_name.startswith("config") and file_name.endswith(".yaml"):
                try:
                    with open(os.path.join(self.program_server_path, file_name), "r", encoding="utf-8") as f:
                        config_data = yaml.safe_load(f)

                    program_server_version = config_data["program_version_number"] # Получаем версию программы на сервере
                    
                    if not program_server_version:
                        return None

                    if program_server_version <= self.program_version_number:
                        return True
                except IndexError:
                    # Если формат имени файла не соответствует ожидаемому, пропускаем его
                    continue
        else:
            return False

    def update_product_names(self):
        """Функция обновляет список названий изделий, рекурсивно обходя папку."""
        products_names = []  # Создаем пустой список названий изделий

        if not self.is_products_folder_available:  # Если папка изделий недоступна
            self.products_names = products_names  # Устанавливаем пустой список названий изделий
            self.show_notification.emit("error", "Папка изделий недоступна на сервере.")
        try:
            for root, dirs, files in os.walk(self.path_to_products_folder):  # Рекурсивно обходим папку изделий
                # Определяем, какие из подпапок являются папками продуктов (т.е. не "рмп")
                product_subdirs = [d for d in dirs if d.lower() != 'рмп']

                # Исключаем папку "рмп" из дальнейшего обхода os.walk
                # Модификация dirs влияет на последующие итерации os.walk
                if 'рмп' in [d.lower() for d in dirs]:
                    dirs.remove('рмп' if 'рмп' in dirs else 'РМП')

                # Если в текущей папке нет вложенных папок-продуктов, то она сама является продуктом
                if not product_subdirs:
                    relative_path = os.path.relpath(
                        root, self.path_to_products_folder)

                    if relative_path != '.':
                        products_names.append(relative_path.split(os.sep))

            self.products_names = products_names
        except Exception as e:
            self.show_notification.emit("error", f"Ошибка при обновлении списка названий изделий: {e}")

    def update_program(self):
        """Функция вызывает обновление программы"""
        updater_path = os.path.join(os.getcwd(), "updater.exe") # Создаём путь к программе обновления
        
        # Проверяем, что updater.exe существует
        if not os.path.exists(updater_path):
            self.show_notification.emit("error", f"Не найден файл обновления: {updater_path}")
            return

        # Запускаем updater.exe с запросом прав администратора
        # Используем 'runas' для повышения прав
        subprocess.run(['powershell', '-Command', f'Start-Process "{updater_path}" -ArgumentList "{self.program_server_path}" -Verb RunAs'], shell=True)


    def export_data(self, save_path, export_format, data):
        """Функция экспортирует данные в указанный формат."""
        if export_format is None:
            self.show_notification.emit("error", "Не выбран формат экспорта.")
            return
        
        if not data:
            self.show_notification.emit("error", "Нет данных для экспорта.")
            return
        
        save_file_path = os.path.join(save_path, f"{self.current_product}.{export_format}")

        if export_format == "xlsx":
            self.__export_to_excel(save_path=save_file_path, data=data)
        elif export_format == "docx":
            self.__export_to_word(save_path=save_file_path, data=data)
        elif export_format == "pdf":
            self.__export_to_pdf(save_path=save_file_path, data=data)
        else:
            self.show_notification.emit("error", "Неподдерживаемый формат экспорта.")
            return