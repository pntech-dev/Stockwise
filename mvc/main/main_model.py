import os
import subprocess
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from PyQt5.QtCore import QObject, pyqtSignal

# Column keys in the source Excel files
NOM_KEY = "Номенклатура"
QTY_KEY = "Количество"
UNIT_KEY = "Ед. изм."
RMP_KEY = "РМП"


class MainModel(QObject):
    """Manages the application's data and business logic.

    This class handles loading configuration, finding and processing product data,
    checking for updates, and exporting data to Excel. It communicates with
    the controller/view using signals.

    Attributes:
        show_notification: A signal that emits messages to be displayed to the user.
    """

    show_notification = pyqtSignal(str, str)

    def __init__(self) -> None:
        """Initializes the model and loads configuration/state."""
        super().__init__()
        self.program_version_number: str = ""
        self.program_server_path: str = ""
        self.path_to_products_folder: str = ""
        self.is_products_folder_available: bool = False
        self._load_config()

        self.products_names: List[List[str]] = []
        self.update_products_names()

        self.current_product: str = ""
        self.current_product_path: str = ""
        self.current_semi_finished_products: List[str] = []
        self.current_product_materials: List[Dict] = []
        self.material_selection: Dict[str, bool] = {}
        self.norms_calculations_value: int = 1
        self.search_in_materials: bool = False
        self.search_in_materials_data: List[Dict] = []

    def get_semi_finished_products(self, product_name: tuple) -> List[str]:
        """Collects Excel files for the selected product.

        Args:
            product_name: Tuple of path segments that identify the product.

        Returns:
            List of Excel file paths, including files inside the RMP subfolder.
        """

        def add_file(file_path: str, file_list: List[str]) -> None:
            if os.path.isfile(file_path) and file_path.lower().endswith((".xlsx", ".xls")):
                file_list.append(file_path)

        semi_finished_products: List[str] = []
        try:
            product_path = os.path.join(self.path_to_products_folder, *product_name)
            if os.path.exists(product_path) and os.path.isdir(product_path):
                self.current_product_path = product_path
                for item_name in os.listdir(product_path):
                    item_path = os.path.join(product_path, item_name)
                    if item_name.lower() == "рмп" and os.path.isdir(item_path):
                        for rmp_file_name in os.listdir(item_path):
                            add_file(os.path.join(item_path, rmp_file_name), semi_finished_products)
                    else:
                        add_file(item_path, semi_finished_products)
            self.current_semi_finished_products = semi_finished_products
            return semi_finished_products
        except Exception as e:
            self.current_semi_finished_products = []
            self.show_notification.emit(
                "error", 
                f"Произошла ошибка в процессе объединения полуфабрикатов {e}"
                )
            
            return []

    def get_product_materials(self, semi_finished_products: List[str]) -> List[Dict]:
        """Aggregates materials from semi-finished product workbooks.

        Args:
            semi_finished_products: Paths to Excel files that describe semi-finished items.

        Returns:
            Deduplicated materials list with merged quantities and RMP markers.
        """
        product_materials_dict: Dict[str, Dict] = {}
        semi_product_names = [
            os.path.splitext(os.path.basename(p))[0].lower().strip() for p in semi_finished_products
        ]

        for file_path in semi_finished_products:
            if not os.path.exists(file_path):
                self.show_notification.emit("error", f"Файл не найден.\nПуть: {file_path}")
                continue

            try:
                is_rmp = os.path.basename(os.path.dirname(file_path)).lower() == "рмп"
                df = pd.read_excel(file_path, sheet_name="TDSheet")
                df = self._normalize_material_columns(df)
                df = df.dropna(subset=[NOM_KEY]).fillna("")

                for item in df.to_dict("records"):
                    nomenclature = item[NOM_KEY]
                    if nomenclature.lower().strip() in semi_product_names:
                        continue

                    quantity = item[QTY_KEY] / 1000
                    if self.norms_calculations_value != 1:
                        quantity *= self.norms_calculations_value
                    item[QTY_KEY] = quantity
                    item[RMP_KEY] = is_rmp

                    if nomenclature in product_materials_dict:
                        product_materials_dict[nomenclature][QTY_KEY] += quantity
                    else:
                        product_materials_dict[nomenclature] = item

            except Exception as e:
                self.show_notification.emit(
                    "error", 
                    f"Произошла ошибка в процессе чтения файла: {os.path.basename(file_path)}: {e}"
                    )

        return list(product_materials_dict.values())

    def recalculate_current_materials(self) -> List[Dict]:
        """Recalculates materials for the current product using the stored semi-finished list."""
        if not self.current_semi_finished_products:
            self.current_product_materials = []
            self.sync_material_selection([], reset=True)
            return []

        product_materials = self.get_product_materials(self.current_semi_finished_products)
        self.current_product_materials = product_materials
        self.sync_material_selection(product_materials, reset=False)
        return product_materials

    def sync_material_selection(self, materials: List[Dict], reset: bool = False) -> None:
        """Syncs selection map with current materials list.

        Args:
            materials: Materials currently displayed.
            reset: If True, default all to selected.
        """
        new_selection: Dict[str, bool] = {}
        for item in materials:
            name = item[NOM_KEY]
            new_selection[name] = True if reset or name not in self.material_selection else self.material_selection[name]
        self.material_selection = new_selection

    def set_material_selected(self, name: str, is_selected: bool) -> None:
        """Sets selection state for a single material.

        Args:
            name: Material name used as the selection key.
            is_selected: Desired selection state.
        """
        if name in self.material_selection:
            self.material_selection[name] = is_selected

    def set_all_materials_selected(self, is_selected: bool) -> None:
        """Sets selection state for all materials.

        Args:
            is_selected: Desired selection state applied to every material.
        """
        for key in list(self.material_selection.keys()):
            self.material_selection[key] = is_selected

    def get_desktop_path(self) -> Optional[Path]:
        """Resolves a Desktop path, handling OneDrive RU/EN variants.

        Returns:
            Desktop path if it can be resolved, otherwise ``None``.
        """
        try:
            home = Path.home()
            onedrive_ru = home / "OneDrive" / "Рабочий стол"
            if onedrive_ru.exists():
                return onedrive_ru
            onedrive_en = home / "OneDrive" / "Desktop"
            if onedrive_en.exists():
                return onedrive_en
            return home / "Desktop"
        
        except Exception as e:
            self.show_notification.emit(
                "error", 
                f"Не удалось определить путь к рабочему столу: {e}"
                )
            
            return None

    def check_program_version(self) -> Optional[bool]:
        """Compares the local version with the server configuration.

        Returns:
            True if up-to-date, False if update needed, None on error/no path.
        """
        if not self.program_server_path or not os.path.exists(self.program_server_path):
            return None

        for file_name in os.listdir(self.program_server_path):
            if file_name.startswith("config") and file_name.endswith(".yaml"):
                try:
                    with open(os.path.join(self.program_server_path, file_name), "r", encoding="utf-8") as f:
                        config_data = yaml.safe_load(f)
                    server_version = config_data.get("program_version_number")
                    if server_version is None:
                        return None
                    return server_version <= self.program_version_number
                except (IOError, yaml.YAMLError, KeyError):
                    continue
        return False

    def update_products_names(self) -> None:
        """Refreshes the products list by scanning the configured folder."""
        if not self.is_products_folder_available:
            self.products_names = []
            self.show_notification.emit("error", "Папка изделий недоступна")
            return

        products: List[List[str]] = []
        try:
            for root, dirs, _ in os.walk(self.path_to_products_folder):
                lower_dirs = [d.lower() for d in dirs]
                if "рмп" in lower_dirs:
                    dirs.remove(dirs[lower_dirs.index("рмп")])
                if not dirs:
                    relative_path = os.path.relpath(root, self.path_to_products_folder)
                    if relative_path != ".":
                        products.append(relative_path.split(os.sep))
            self.products_names = products
        except Exception as e:
            self.show_notification.emit(
                "error", 
                f"Произошла ошибка в процесее сканирования продуктов{e}"
                )

    def update_program(self) -> None:
        """Launches the updater executable with admin rights."""
        updater_path = os.path.join(os.getcwd(), "updater.exe")
        if not os.path.exists(updater_path):
            self.show_notification.emit(
                "error", 
                f"Программа автоматического обновления не найдена.\nПуть: {updater_path}"
                )
            
            return
        
        try:
            command = f'Start-Process "{updater_path}" -ArgumentList "{self.program_server_path}" -Verb RunAs'
            subprocess.run(["powershell", "-Command", command], check=True, shell=True)
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            self.show_notification.emit(
                "error", 
                f"Во время запуска программы автоматического удаления произошла ошибка: {e}"
                )

    def open_config_file(self) -> None:
        """Launches the config file"""
        config_path = os.path.join(os.getcwd(), "config.yaml")
        if not os.path.exists(config_path):
            self.show_notification.emit("error", f"Файл конфигурации не найден.\nПуть: {config_path}")
            return
        
        try:
            os.startfile(config_path)
        except Exception as e:
            self.show_notification.emit(
                "error", 
                f"Произошла ошибка в процессе открытия файла конфигурации: {e}"
                )

    def export_data(self) -> None:
        """Exports selected materials to Excel using the table template.

        The method applies selection state, writes rows into the template, and
        saves the result to the user's Desktop.
        """
        try:
            workbook = load_workbook("templates/table.xlsx")
            sheet = workbook.active
            sheet.title = self._sanitize_sheet_title("Материалы")

            selected_data = [
                item for item in self.current_product_materials if self.material_selection.get(item[NOM_KEY], True)
            ]
            if not selected_data:
                self.show_notification.emit("warning", "Нет данных для экспорта.")
                return

            start_row = 2
            for i, item in enumerate(selected_data):
                row_idx = start_row + i
                sheet.cell(row=row_idx, column=1, value=item[NOM_KEY])
                sheet.cell(row=row_idx, column=2, value=item[UNIT_KEY])
                sheet.cell(row=row_idx, column=3, value=item[QTY_KEY])

            font = Font(name="Times New Roman", size=14)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
            )

            for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, max_col=3):
                for cell in row:
                    cell.font = font
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                sheet.cell(row=cell.row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                sheet.cell(row=cell.row, column=2).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                sheet.cell(row=cell.row, column=3).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0

            desktop_path = self.get_desktop_path()
            if not desktop_path:
                self.show_notification.emit("error", "Не удалось определить путь сохранения.")
                return

            safe_product = self._sanitize_filename(self.current_product) or ""
            file_name = f"Материалы изделия {safe_product}.xlsx"
            workbook.save(desktop_path / file_name)
            self.show_notification.emit("info", f"Экспорт выполнен: {file_name}")

        except Exception as e:
            self.show_notification.emit("error", f"Не удалось выполнить экспорт: {e}")
        finally:
            self.show_notification.emit("", "")

    def _sanitize_sheet_title(self, title: str) -> str:
        """Removes invalid characters from an Excel sheet title.

        Args:
            title: Proposed worksheet name.

        Returns:
            Safe worksheet title compatible with Excel.
        """
        invalid_chars = set(r'[]:*?/\\')
        sanitized = "".join(ch for ch in title if ch not in invalid_chars)
        return sanitized or "Sheet1"

    def _sanitize_filename(self, name: str) -> str:
        """Removes invalid characters from filename components.

        Args:
            name: Raw filename fragment.

        Returns:
            Safe filename fragment; defaults to ``"file"`` when empty.
        """
        invalid_chars = set('\\/:*?"<>|')
        cleaned = "".join(ch for ch in name if ch not in invalid_chars).strip()
        return cleaned or "file"

    def _load_config(self) -> None:
        """Loads configuration from config.yaml and sets product paths."""
        try:
            with open("config.yaml", "r", encoding="utf-8") as file:
                config = yaml.safe_load(file)
        except FileNotFoundError:
            self.show_notification.emit("error", "Файл config.yaml не найден.")
            return

        if not isinstance(config, dict) or "path_to_products_folder" not in config:
            self.show_notification.emit("error", "Файл config.yaml не содержит обязательных полей.")
            return

        self.program_version_number = config.get("program_version_number", "")
        self.program_server_path = config.get("server_program_path", "")

        path = config["path_to_products_folder"]
        if os.path.exists(path) and os.path.isdir(path):
            self.path_to_products_folder = path
            self.is_products_folder_available = True
        else:
            self.show_notification.emit("error", "Путь к папке с продуктами недоступен.")
            self.is_products_folder_available = False

    def _normalize_material_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Renames material columns to standard keys using substring matching.

        Args:
            df: Raw DataFrame read from a materials workbook.

        Returns:
            DataFrame with normalized column names.

        Raises:
            ValueError: If required columns cannot be located.
        """
        columns_lower = {col: str(col).lower() for col in df.columns}

        def find_column(substrings: List[str]) -> Optional[str]:
            for col, low in columns_lower.items():
                if all(sub in low for sub in substrings):
                    return col
            return None

        mapping: Dict[str, str] = {}
        candidates = {
            NOM_KEY: ["номенк"],
            QTY_KEY: ["кол"],
            UNIT_KEY: ["ед", "изм"],
        }

        for target, substrings in candidates.items():
            col = find_column(substrings)
            if col:
                mapping[col] = target

        df = df.rename(columns=mapping)

        missing = [key for key in (NOM_KEY, QTY_KEY, UNIT_KEY) if key not in df.columns]
        if missing:
            raise ValueError(f"Usecols do not match columns, missing: {missing}, got: {list(df.columns)}")

        return df
